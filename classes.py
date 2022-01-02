import datetime
import time
import pyinputplus
import openpyxl
# <editor-fold desc="access exell">
wb_name = 'words collection.xlsx'
bank_deck_name = 'bank_deck'
working_deck_name = 'working_deck'
words_collection_wb = openpyxl.open(wb_name)
# </editor-fold>


class BankDeck:
    def __init__(self, deck_name='bank_deck'):
        self.deck_name = deck_name
        self.deck = {}
        deck_sheet = words_collection_wb[self.deck_name]
        for row in range(2, deck_sheet.max_row+1):
            key = deck_sheet.cell(row=row, column=1).value
            word = deck_sheet.cell(row=row, column=2).value
            self.deck[key] = word

    def clear_deck(self):
        for sheet in words_collection_wb.sheetnames:
            if sheet.startswith(self.deck_name):
                del words_collection_wb[self.deck_name]
        words_collection_wb.create_sheet(self.deck_name)
        deck_sheet = words_collection_wb[self.deck_name]
        deck_sheet['A1'] = 'key'
        deck_sheet['B1'] = 'word'
        words_collection_wb.save(wb_name)
        self.deck = {}

    def get_deck(self):
        return self.deck

    def add_new_word(self, word):
        if word not in self.deck.values():
            deck_sheet = words_collection_wb[self.deck_name]
            time_stamp = datetime.datetime.now()
            deck_sheet.cell(row=deck_sheet.max_row+1, column=1).value = time_stamp
            deck_sheet.cell(row=deck_sheet.max_row, column=2).value = word
            words_collection_wb.save(wb_name)
            self.deck.setdefault(time_stamp, word)
            time.sleep(0.1)
            print(word, 'goes into', self.deck_name)
        else:
            print(f'{word} already in {self.deck_name}')

    def print_words(self):
        print(self.deck_name, end=': ')
        if len(self.deck) == 0:
            print(f'{self.deck_name} is empty')
        else:
            lis = list(self.deck.values())
            for value in lis[0:-1]:
                print(value, end=' ')
            print(lis[-1])


class WorkingDeck(BankDeck):
    def __init__(self):
        super().__init__(deck_name='working_deck')
        self.bank_deck = BankDeck()
        self.learned_deck = {}

    def pick_word_from_bank_deck(self):
        oldest_word_time = datetime.datetime.now()
        for word_time_stamp_key in self.bank_deck.deck:
            if word_time_stamp_key is not None:
                if oldest_word_time >= word_time_stamp_key:
                    oldest_word_time = word_time_stamp_key
        print('oldest_word_time', oldest_word_time)
        for k, v in self.bank_deck.deck.items():
            print(k, v)

        # know_it = 'don*t know'
        know_it = pyinputplus.inputMenu(['know', 'don*t know'], f'know {self.bank_deck.deck[oldest_word_time]}: \n', numbered=True)
        if know_it == 'know':
            self.bank_deck.print_words()
            self.print_words()
            print(f'ah, u know "{self.bank_deck.deck[oldest_word_time]}"')
            self.bank_deck.deck[datetime.datetime.now()] = self.bank_deck.deck[oldest_word_time]
            del self.bank_deck.deck[oldest_word_time]
            self.bank_deck.print_words()
            self.print_words()

        else:
            # <editor-fold desc="sheet adjustment">
            bank_deck_sheet = words_collection_wb[self.bank_deck.deck_name]
            working_deck_sheet = words_collection_wb[self.deck_name]
            for row in range(2, bank_deck_sheet.max_row):
                if bank_deck_sheet.cell(row=row, column=2).value is not None:
                    if bank_deck_sheet.cell(row=row, column=2).value == self.bank_deck.deck[oldest_word_time]:
                        working_deck_sheet.cell(row=working_deck_sheet.max_row + 1, column=1).value = datetime.datetime.now()
                        working_deck_sheet.cell(row=working_deck_sheet.max_row, column=2).value = self.bank_deck.deck[oldest_word_time]
                        bank_deck_sheet.delete_rows(row, 1)
                        words_collection_wb.save(wb_name)
            # </editor-fold>

            # <editor-fold desc="deck dict adjustment">
            super().__init__(deck_name=self.deck_name)
            print(f'{self.bank_deck.deck[oldest_word_time]} ==> moved from {self.bank_deck.deck_name} ==> {self.deck_name}')
            del self.bank_deck.deck[oldest_word_time]
            super().__init__(deck_name=self.bank_deck.deck_name)

            # </editor-fold>

    def move_words_to_bank_deck(self):
        print(self.deck)
        print(self.bank_deck)
        for key in self.deck.keys():
            self.bank_deck.deck[datetime.datetime.now()] = self.deck[key]
        self.deck = {}
        words_collection_wb.save(wb_name) # here we saving the sheet but we have not made any changes to it,
        # meanwhile we should, like move words in wb from 1 sheet to another, and only then save it
        print(f'we moved words from {self.deck_name} to bank_deck')
        print(self.deck)
        print(self.bank_deck)

    def rotation(self):
        while len(self.deck) > 0:
            for key in list(self.deck.keys()):
                remember = 'yes'  # pyinputplus.inputMenu(['yes', 'no'], f'know {self.deck[key]} ?\n', numbered=True)
                if remember == 'yes':
                    self.learned_deck[datetime.datetime.now()] = self.deck[key]
                    print(f'you have learned {self.deck[key]} so it goes out of this rotation')

                    del self.deck[key]
                if remember == 'no':
                    print(f'{self.deck[key]} goes in the end of the {self.deck_name}')
        print(f'you have repeated all words from {self.deck_name}')
        leave_words = pyinputplus.inputMenu(['leave', 'move'], f'leave words in {self.deck_name} or move to bank_deck?\n', numbered=True)
        if leave_words == 'leave':
            pass
            print('as u wish')
        if leave_words == 'move':
            self.move_words_to_bank_deck()
